import yaml
import openpyxl
from openpyxl.styles import Font
from paramiko import SSHClient, AutoAddPolicy
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import re
import ipaddress


class Server:
    def __init__(self, identifier, username="smnair"):
        self.identifier = identifier
        self.username = username
        self.ssh = SSHClient()
        self.ssh.set_missing_host_key_policy(AutoAddPolicy())
        self.ssh.connect(self.identifier, username=self.username)

    def exec_command(self, command):
        stdin, stdout, stderr = self.ssh.exec_command(command)
        return stdout.read().decode().strip()

    def close(self):
        self.ssh.close()


def load_servers(yaml_file):
    with open(yaml_file, "r") as file:
        config = yaml.safe_load(file)
    return config.get("servers", []) + config.get("ip_servers", [])


def is_valid_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


def get_server_info(server_identifier):
    print(f"Connecting to {server_identifier}...")

    try:
        server = Server(server_identifier)

        # Get hostname
        hostname = server.exec_command("hostname")

        # GPU Info
        gpu_info = server.exec_command(
            "nvidia-smi --query-gpu=name,memory.total --format=csv,noheader"
        )
        gpus = []
        if gpu_info:
            for line in gpu_info.split("\n"):
                if line.strip():
                    name, vram = line.split(",")
                    gpus.append({"type": name.strip(), "vram": vram.strip()})

        # CPU Info
        cpu_count = server.exec_command("nproc")
        cpu_type = server.exec_command("lscpu | grep 'Model name' | cut -f 2 -d ':'")

        # RAM Info
        ram_info = server.exec_command("free -h | awk '/^Mem:/ {print $2}'")

        # Storage Info
        df_output = server.exec_command("df -h")
        storage_info = []
        for line in df_output.split("\n")[1:]:  # Skip header
            parts = line.split()
            if len(parts) >= 6:
                device, size, used, available, use_percent, mount = parts[:6]
                # Filter out loop devices, small partitions, and non-standard mounts
                if (
                    device.startswith("/dev/")
                    and not device.startswith("/dev/loop")
                    and not re.match(r"/dev/.*p\d+$", device)
                    and not mount.startswith("/snap/")  # Exclude partitions
                    and not mount.startswith("/boot/")
                    and size.endswith(("T", "G"))
                    and not size.startswith(  # Only include TB or GB sizes
                        ("1G", "2G", "3G", "4G", "5G")
                    )
                ):  # Exclude smaller sizes
                    uuid = server.exec_command(f"sudo blkid -s UUID -o value {device}")
                    storage_info.append(
                        {
                            "device": device,
                            "mount": mount,
                            "uuid": uuid.strip(),
                            "total": size,
                            "available": available,
                        }
                    )

        server.close()

        return server_identifier, {
            "hostname": hostname,
            "gpu": gpus,
            "cpu": {"count": cpu_count.strip(), "type": cpu_type.strip()},
            "ram": ram_info.strip(),
            "storage": storage_info,
        }
    except Exception as e:
        print(f"Error collecting data from {server_identifier}: {str(e)}")
        return server_identifier, None


def create_spreadsheet(servers_data, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Server Inventory"

    headers = [
        "Server Identifier",
        "Hostname",
        "GPU Type",
        "GPU VRAM",
        "CPU Count",
        "CPU Type",
        "RAM",
        "Storage Device",
        "Mount Path",
        "UUID",
        "Total Storage",
        "Available Storage",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    row = 2
    for server_id, data in servers_data.items():
        if data is None:
            ws.cell(row=row, column=1, value=server_id)
            ws.cell(row=row, column=2, value="Error collecting data")
            row += 1
            continue

        hostname = data.get("hostname", "")
        gpu_info = data.get("gpu", [{}])
        cpu_info = data.get("cpu", {})
        ram_info = data.get("ram", "")
        storage_info = data.get("storage", [])

        # GPU info (may be multiple)
        for gpu in gpu_info:
            ws.cell(row=row, column=1, value=server_id)
            ws.cell(row=row, column=2, value=hostname)
            ws.cell(row=row, column=3, value=gpu.get("type", ""))
            ws.cell(row=row, column=4, value=gpu.get("vram", ""))
            row += 1

        # If no GPU, still add server info
        if not gpu_info:
            ws.cell(row=row, column=1, value=server_id)
            ws.cell(row=row, column=2, value=hostname)
            row += 1

        # CPU info
        ws.cell(row=row - 1, column=5, value=cpu_info.get("count", ""))
        ws.cell(row=row - 1, column=6, value=cpu_info.get("type", ""))

        # RAM info
        ws.cell(row=row - 1, column=7, value=ram_info)

        # Storage info (may be multiple)
        for storage in storage_info:
            ws.cell(row=row - 1, column=8, value=storage.get("device", ""))
            ws.cell(row=row - 1, column=9, value=storage.get("mount", ""))
            ws.cell(row=row - 1, column=10, value=storage.get("uuid", ""))
            ws.cell(row=row - 1, column=11, value=storage.get("total", ""))
            ws.cell(row=row - 1, column=12, value=storage.get("available", ""))
            row += 1

    wb.save(output_file)


def main():
    start_time = time.time()
    servers_list = load_servers("servers.yaml")
    servers_data = {}

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_server = {
            executor.submit(get_server_info, server): server for server in servers_list
        }
        for future in as_completed(future_to_server):
            server_identifier, server_data = future.result()
            if server_data:
                servers_data[server_identifier] = server_data

    create_spreadsheet(servers_data, "uwaterloo_server_inventory.xlsx")
    print(f"Inventory spreadsheet created: uwaterloo_server_inventory.xlsx")
    print(f"Total execution time: {time.time() - start_time:.2f} seconds")


if __name__ == "__main__":
    main()
